# workshop8 task3 - load and explore an Excel data file, 17/09/21, Cordelia Jiang
# import openpyxl modules.
import openpyxl

# get the IBM worksheet of the Stocks.xlsx workbook
wb = openpyxl.load_workbook('Stocks.xlsx')
sheet = wb['IBM']

# retrieve all data in 2010.
# go through all rows in sheet, check if first column has 2010 as its value of year.
# print value of cells for each rows only if its first column has 2010 as its year.
for row in sheet.iter_rows(min_row=1, min_col=1, max_row=sheet.max_row, max_col=sheet.max_column):
    if row[0].value.year == 2010:
        for cell in row:
            print(cell.value, end=" ")
        print()

# a variable used to store 2010 data as list items
data2010 = []

# calculate min, max, mean of 2010 data
# go through all rows in sheet, check if first column has 2010 as its value of year.
# append value only if its first column has 2010 as its year.
for row in sheet.iter_rows(min_row=1, min_col=1, max_row=sheet.max_row, max_col=sheet.max_column):
    if row[0].value.year == 2010:
        data2010.append(row[3].value)

# print min, max and mean of 2010 data, each printing statement will start on a new line
print("Min of 2010 data: ", min(data2010), end="\n")
print("Max of 2010 data: ", max(data2010), end="\n")
print("Mean of 2010 data: ", sum(data2010)/len(data2010), end="\n")


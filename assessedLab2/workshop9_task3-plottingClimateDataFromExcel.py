# workshop9 task3 - plotting climate data from excel, 21/09/21, Cordelia Jiang
# use the openpyxl module to load the ClimateData.xlsx workbook
import openpyxl
import pandas as pd    # used for calculate the average for each year and store it for each city
import matplotlib.pyplot as plt    # import needed module to perform plot functions

# get the Sheet1 worksheet of the ClimateData.xlsx workbook
wb = openpyxl.load_workbook('ClimateData.xlsx')
sheet = wb['Sheet1']

# variables used to store each city's year and average temperature data as list items
sydneyData = []
londonData = []
newYorkData = []
jakartaData = []

# min_row needs to start from 2, as the first row is the data titles
for row in sheet.iter_rows(min_row=2, min_col=1, max_row=sheet.max_row, max_col=sheet.max_column):
    x = 1900
    # remove all empty rows in sheet before processing rows, only proceed the for loop if there is no empty rows
    if row[0] is not None:
        # check if year is between 1900 and 2000 (inclusive)
        if 1900 <= row[0].value.year <= 2000:    # same as row[0].value.year >= 1900 and row[0].value.year <= 2000
            # store year and average temperature for Sydney, London, New York, and Jakarta each in a separated list
            if row[3].value == "Sydney":
                sydneyData.append([row[0].value.year, row[1].value])
            if row[3].value == "London":
                londonData.append([row[0].value.year, row[1].value])
            if row[3].value == "New York":
                newYorkData.append([row[0].value.year, row[1].value])
            if row[3].value == "Jakarta":
                jakartaData.append([row[0].value.year, row[1].value])
    else:
        continue    # continue with the next iteration

# calculate the average temperature for each year of sydney using pandas
sydneyYearPlot = []    # empty list used to store and plot year for sydney
sydneyAvgTempPlot = []    # empty list used to store and plot average temperature for sydney
columns = ["Sydney Year", "Sydney Avg Temperature"]
df = pd.DataFrame(sydneyData, columns=columns)    # takes sydneyData as its data, and set items in columns as its labels

# group sydneyData by its year, and get the mean of sydney's average temperature
meanResult = df.groupby(["Sydney Year"])["Sydney Avg Temperature"].mean()

# go through items in meanResult, store sydneyYear in sydneyYearPlot for plotting,
# and store sydneyAvgTemp in sydneyAvgTempPlot for plotting
for sydneyYear, sydneyAvgTemp in meanResult.items():
    sydneyYearPlot.append(sydneyYear)
    sydneyAvgTempPlot.append(sydneyAvgTemp)

# calculate the average temperature for each year of london using pandas
londonYearPlot = []    # empty list used to store and plot year for london
londonAvgTempPlot = []    # empty list used to store and plot average temperature for london
columns = ["London Year", "London Avg Temperature"]
df = pd.DataFrame(londonData, columns=columns)    # takes londonData as its data, and set items in columns as its labels

# group londonData by its year, and get the mean of london's average temperature
meanResult = df.groupby(["London Year"])["London Avg Temperature"].mean()

# go through items in meanResult, store londonYear in londonYearPlot for plotting,
# and store londonAvgTemp in londonAvgTempPlot for plotting
for londonYear, londonAvgTemp in meanResult.items():
    londonYearPlot.append(londonYear)
    londonAvgTempPlot.append(londonAvgTemp)

# calculate the average temperature for each year of new york using pandas
newYorkYearPlot = []    # empty list used to store and plot year for new york
newYorkAvgTempPlot = []    # empty list used to store and plot average temperature for new york
columns = ["New York Year", "New York Avg Temperature"]
df = pd.DataFrame(newYorkData, columns=columns)  # takes newYorkData as its data, and set items in columns as its labels

# group newYorkData by its year, and get the mean of new york's average temperature
meanResult = df.groupby(["New York Year"])["New York Avg Temperature"].mean()

# go through items in meanResult, store newYorkYear in newYorkYearPlot for plotting,
# and store newYorkAvgTemp in newYorkAvgTempPlot for plotting
for newYorkYear, newYorkAvgTemp in meanResult.items():
    newYorkYearPlot.append(newYorkYear)
    newYorkAvgTempPlot.append(newYorkAvgTemp)

# calculate the average temperature for each year of jakarta using pandas
jakartaYearPlot = []    # empty list used to store and plot year for jakarta
jakartaAvgTempPlot = []    # empty list used to store and plot average temperature for jakarta
columns = ["Jakarta Year", "Jakarta Avg Temperature"]
df = pd.DataFrame(jakartaData, columns=columns)  # takes jakartaData as its data, and set items in columns as its labels

# group jakartaData by its year, and get the mean of jakarta's average temperature
meanResult = df.groupby(["Jakarta Year"])["Jakarta Avg Temperature"].mean()

# go through items in meanResult, store jakartaYear in jakartaYearPlot for plotting,
# and store jakartaAvgTemp in jakartaAvgTempPlot for plotting
for jakartaYear, jakartaAvgTemp in meanResult.items():
    jakartaYearPlot.append(jakartaYear)
    jakartaAvgTempPlot.append(jakartaAvgTemp)

# generate a line graph using matplotlib to compare the average yearly temperatures for Sydney, London,
# New York and Jakarta for the 20th century (1900-2000 inclusive).
plt.plot(sydneyYearPlot, sydneyAvgTempPlot, londonYearPlot, londonAvgTempPlot, newYorkYearPlot, newYorkAvgTempPlot,
         jakartaYearPlot, jakartaAvgTempPlot)

# include the relevant title, xy axis labels and data labels using the legend() function.
plt.title('Year Versus Average Temperature')
plt.xlabel('Year')
plt.ylabel('Average Temperature')
plt.legend(["Sydney", "London", "New York", "Jakarta"], loc="best")    # put legend for cities at the "best" location

# display line graph
plt.show()
